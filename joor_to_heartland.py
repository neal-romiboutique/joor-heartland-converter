#!/usr/bin/env python3
"""
JOOR -> Heartland Retail converter for Romi Boutique.  (v7)

Phase 1 (grids):  python3 joor_to_heartland.py grids JOOR_ORDER.csv|.xlsx [-o grids.xlsx]
    One row per (vendor, style, color) in Heartland's official grid-import
    template schema, Generate Items = true, sizes completed per store rules.

Phase 2 (po):     python3 joor_to_heartland.py po JOOR_ORDER.csv|.xlsx ITEMS_EXPORT.csv [-o po.xlsx]
    Joins Heartland's post-import grid/items export back to the JOOR order
    (grid # + size -> Item #) and emits the PO import file.

Accepts BOTH JOOR export formats, auto-detected:
  A) Basic Export (long): csv/tsv/txt/xlsx, header row 1, snake_case columns,
     one row per style/color/size.
  B) Order-to-Size (wide): formatted xlsx PO, metadata block on top, one row
     per style/color with sizes as columns; may contain embedded product
     images, which the grids phase extracts to <output>_images/.

Store conventions encoded below; edit CONFIG to change.
"""
import argparse, csv, math, os, re, sys, unicodedata
import requests
from decimal import Decimal, ROUND_CEILING
from pathlib import Path

# ----------------------------- CONFIG ---------------------------------------
MARGIN = Decimal("2.4")          # retail = MARGIN x unit cost when JOOR has no MSRP
VENDOR_MAP = {                   # raw JOOR seller_account -> Heartland vendor record
    # "Some JOOR Name": "Exact Heartland Vendor Name",  # explicit overrides go here
}
# Known Heartland vendors (Purchasing -> Vendors export), used to resolve the
# exact existing vendor Name instead of guessing one. Re-export and regenerate
# this list whenever new vendors are added in Heartland.
KNOWN_VENDOR_NAMES = [
    '143',
    '27 Miles',
    '3 X 1',
    '360 Cashmere',
    '382',
    '525',
    '7 For All Mankind',
    '9SEED',
    'A Shirt Thing',
    'A.L.C',
    'AGOLDE',
    'AMERICAN COLORS',
    'APPARIS',
    'AS 98',
    'AS BY DF',
    'AS98',
    'ASKK NY',
    'ATM',
    'Abigail Jayne Design',
    'Adina Reyter',
    'Alemais',
    'Alexa Leigh',
    'Alice & Olivia',
    'Alice & Trixie',
    'Alix of Bohemia',
    'All Time Favorite',
    'Aloha',
    'Amalia',
    'Amanda Uprichard',
    'American Vintage',
    'Amo',
    'Ancient Greek Sandals',
    'Angel Wings',
    'Anonym',
    'Ansaldo',
    'Apres Jewelry',
    'Aquarius Cocktail',
    'Aquatalia',
    'Autumn Cashmere',
    'Avenue Montaigne',
    'Aviator Nation',
    'Azul',
    'B. Belt',
    'B.Belt',
    'BA&SH',
    'BALLIER',
    'BB Dakota',
    'BE JE',
    'BLOOM JEWELRY',
    'BOHO',
    'Banjanan',
    'Be Love',
    'Beek',
    'Bell',
    'Bella Dahl',
    'Bella Sleep + Spa',
    'Bevy Flog',
    'Bibi Lou',
    'Blonde No 8',
    'Brazeau Tricot',
    'Britta',
    'Brochu Walker',
    'Brodie',
    'Bronx And Banco',
    'Bubish',
    'CALLAHAN',
    'CALYPSO',
    'CARDEN AVENUE JEWELRY',
    'CATHERINE GEE',
    'CHINTI AND PARKER',
    'CHRISTY LYNN',
    'CINO',
    'CIRCOLO',
    'COLORI',
    'CRUSH',
    'Caballero',
    'Caddis',
    'California Tailor',
    'Calleen Cordero',
    'Cami',
    'Cami NYC',
    'Campomaggi',
    'Cara Cara Clothing',
    'Carrie Forbes',
    'Caterina Lucchi',
    'Central Park WEST',
    'Chan Luu',
    'Charo Ruiz Ibiza',
    'Chaser',
    'Chie Mihara',
    'Christopher Fischer',
    'Ciao Lucia',
    'Cinq A Sept',
    'Citizens',
    'Clare V',
    'Clare Vivier',
    'Claris Virot',
    'Cleobella',
    'Coclico',
    'Colorush',
    'Commando',
    'Corazon Playero',
    'Cordani',
    'Cosabella',
    'Cotton Citizen',
    'Cowboy Belts',
    'Cupcakes & Cashmere',
    'Current Elliot',
    'DANIEL RAHIM',
    'DELACY',
    'DL1961',
    'DOEN',
    'DONNI.',
    'Dear Frances',
    'Denimist',
    'Derek Lam',
    'Design History',
    'Devon Dowd',
    'Diane Von Furstenberg',
    'Dolce Vita',
    'Doma',
    'Dragon Diffusion',
    'Drew',
    'EF Collection',
    'EJ Jewelry',
    'ERIN GRAY',
    'Eberjey',
    'Ecru',
    'Elaine Kim',
    'Electric & Rose',
    'Elizabeth and James',
    'Ella Moss',
    'Embrazio',
    'Enza Costa',
    'Equipment',
    'Estheme Cashmere',
    'Eterne',
    'Eugenia Kim',
    'FARM Rio',
    'FIGUE',
    'FITE',
    'FRANK & EILEEN',
    'FRANK & EILEEN TEE LAB',
    'FREECITY',
    'FUN SOCKS',
    'Faliero Sarti',
    'Feel The Piece',
    'Felicite',
    'Fig & Bella',
    'Florabella',
    'Frame',
    'Freda Salvador',
    'Freya',
    'G1',
    'GANNI',
    'GBTSO',
    'GILNER FARRAR',
    'GOGO',
    'GRLFRND',
    'Galadriel Mattei',
    'Generation Love',
    'Gentle Fawn',
    'Giles and Brothers',
    'Global Girls',
    'Go Silk',
    'Goddis',
    'Gold Hawk',
    'Goldsign',
    'Good hYouman',
    'Groceries Apparel',
    'Guest In Residence',
    'Gypsy',
    'HAPPY SHEEP',
    'HASBEENS',
    'Halston Heritage',
    'Hammitt',
    'Hartford',
    'Haute Hippie',
    'Havaiana',
    'Helene',
    'Hemant & Nandita',
    'Hudson',
    'Hula Hoops',
    "I Stole My Boyfriend's Shirt",
    'ICONS',
    'IKKS',
    'IRO',
    'Inhabit',
    'Isabel Marant',
    'J-ALL',
    'JBrand',
    'JOHN ELLIOTT',
    'Jakett',
    'Jakett New York',
    'James Perse',
    'Jane Hollinger',
    'Jane Tran',
    'Janessa Leone',
    'Jen R',
    'Jennifer Haley',
    'Jennifer Zeuner',
    'Jenny Bird',
    'Jet',
    'Joelle Casady',
    'John & Jen',
    'John & Jenn',
    'Joie',
    'Joy Dravecky',
    'Joy Li',
    'Julie Shenkman',
    'Jumper 1234',
    'June',
    'June 7.2',
    'Junk Food',
    'Jurate Brown',
    'KERO',
    'KERRI ROSENTHAL',
    'KMJ',
    'KNTIWIT',
    'Kane',
    'Kelly Wearstler',
    'Kersh',
    'Kim Vogel',
    'Kim White',
    'Kobo',
    'Krisa',
    'Kujten',
    "L'Agence",
    'LAFCO',
    'LBLC',
    'LE SUPERBE',
    'LNA',
    'LOLA & SOPHIE',
    'LOVE SAM',
    'LOVE TANJANE',
    'La Made',
    'La Natura',
    'La Prestic Ouiston',
    'Lana Bilzerian',
    'Lanston',
    'Lauren & Gracia',
    'Lauren Moshi',
    'Lavender Crush',
    'Le Bon Shoppe',
    'Le Divina',
    'Le Jean',
    'Leallo',
    'Leighelena',
    'Letarte',
    'Level 9',
    'Lilla P',
    'Line',
    'Lingua Franca',
    'Lisa Freede',
    'Lisa Yang',
    'Local Beach',
    'Loeffler Randall',
    'Lola Cruz',
    'Love & Bambii',
    'Love The Label',
    'LoveShackFancy',
    'Loyd/Ford',
    'Lumilla',
    'Luv AJ',
    'Lynn Tallerico',
    'MARIE OLIVER',
    'MARIET MENTAAL',
    'ME & K',
    'MES DEMOISELLES',
    'MILLY CABANA',
    'MOLLY G',
    'Mabe',
    'Madeleine Thompson',
    'Madeworn',
    'Maiami',
    'Maison Hotel',
    'Maison Montagut',
    'Marcia Moran',
    'Marie Chavez',
    'Marit Rae',
    "Mason's",
    'Matta',
    'Mauritius Leather',
    'Maya Bauman',
    'McGuire',
    'Me369',
    'Mercado Global',
    'Mexicana',
    'Michael Lauren',
    'Michael Stars',
    'Mikoh',
    'Milly',
    'Minnie Rose',
    'Mint & Rose',
    'Misa',
    'Miz Mooz',
    'Momoni',
    'Monrow',
    'Mother',
    'Mou',
    'Moussy',
    'Muuba',
    'NICKHO REY',
    'Naadam',
    'Naghedi',
    'Naom',
    'Nation',
    'Native Gem',
    'Never A Wallflower',
    'Nibi MTK',
    'Nili Lotan',
    'Nothing Personal',
    'Nsf',
    'Nubra',
    'ONE GREY DAY',
    'Oats Cashmere',
    'Odette',
    'Olivia Dar',
    'Oncept',
    'Oxbow',
    'PARKER NY',
    'PETE',
    'PHILANTHROPY',
    'PINE CASHMERE',
    'PJ SALVAGE',
    'Paige',
    'Pam & Gela',
    'Paradigm',
    'Parrish',
    'Paul Green',
    'Paula Carvalho',
    'Paula Rosen',
    'Peace Love World',
    'Pharaoh',
    'Pliers & String',
    'Poupette St. Barth',
    'Puntovita',
    'Q',
    'Quilted Koala',
    'RACHEL PALLY',
    'RAQUEL ALLEGRA',
    'ROCK YOUR RADIANCE',
    'ROI',
    'ROSE CARMINE',
    'Rachel Gilbert',
    'Rag & Bone',
    'Rails',
    'Ramona & Ruth',
    'Ramy Brook',
    'Re/Done',
    'Rebecca Scott',
    'Rebels',
    'Retrofete',
    'Richiami',
    'Roam',
    'Robin Haley',
    'Romi',
    'SACCHARINE NY',
    'SALONI',
    'SALTWATER LUXE',
    'SANS FAFF',
    'SANTORE JEWELRY',
    'SARAH DUNN',
    'SEYCHELLES',
    'SOFTWAVES',
    'SOH',
    'SOLID & STRIPED',
    'SPRWMN',
    'STATESIDE',
    'Sablyn',
    'Salt Water LUXE',
    'Sanctuary',
    'Satya Jewelry',
    'Schoen by Yu',
    'Sea',
    'Simon Miller',
    'Skull Cashmere',
    'Soft Joie',
    'Soia & Kyo',
    'Sol Los Angeles',
    'Sold Out NYC',
    'Solei Sea',
    'Sonia & Kyo',
    'Soulbyrd',
    'Splendid',
    'Staud',
    'Streets Ahead',
    'Studio M',
    'Subtle Luxury',
    'Sundry',
    'Susana Monaco',
    'Suzie Kondi',
    'TAI',
    'TKees',
    'TRINA TURK',
    'TWP',
    'TYLHO',
    'Temptation Positano',
    'The Jacksons',
    'Theia',
    'Tilo',
    'Toast',
    'Travata',
    'Tyler Jacobs',
    'Tysa',
    'Ulla Johnson',
    'Unsweetened NY',
    'VICTORIA CUNNINGHAM',
    'Van Palma',
    'Vanzetti',
    'Veronica Beard',
    'Viereck',
    'Vince',
    'Vipul Shah',
    'Vita Fede',
    'Voyage et Cie',
    'W.Cashmere',
    'WESRF',
    'Wandler',
    'White & Warren',
    'Wilt',
    'Wooden Ships',
    'Wyeth',
    'YFB',
    'Yigal Azrouel',
    'Yosi Samra',
    'Yuzefi',
    'Zadig & Voltaire',
    'Zoe Chicco',
    'Zofia Day',
    'ashley ashoff',
    'm&m bling',
    'r & r',
    'rjp',
    'road 22',
    'signorelli',
    'swone',
    'tangerine',
]
SEASON_FROM = "linesheet"        # JOOR column used for Season (long format only)

# Controlled-vocabulary fields on the grid template: rendered as in-cell
# dropdowns so the human reviewer picks from these instead of typing free text.
DROPDOWN_LISTS = {
    "Shopify Category": ["Apparel", "Bags", "Shoes", "Jewelry", "Accessories"],
    "Shopify Color": ["White", "Red", "Burgundy", "Pink", "Orange", "Yellow", "Green",
                      "Blue", "Navy", "Ivory", "Cream", "Tan", "Grey", "Brown", "Black",
                      "Gold", "Multi", "Clear", "Silver", "Yellow Gold", "White Gold",
                      "Two Tone", "Black Diamond", "Emerald", "Turquoise", "Rose Gold",
                      "Pearl", "Labradorite"],
    "Active fit guide": ["1", "2", "3", "4", "5"],
    "Category": ["Dresses", "Denim", "Tops", "T-Shirts", "Sweaters", "Loungewear",
                "Skirts", "Jackets", "Pants", "Shorts", "Handbags", "Shoes", "Scarves",
                "Belts", "Jewelry", "Fine jewelry", "Lingerie", "Sleepwear", "Fragrance",
                "Other Accessories", "Greeting Cards", "Hats & Gloves", "Other"],
}
# Heartland official grid-import template headers (84 cols), verbatim:
TEMPLATE_HEADERS = ['Item Grid #','Description','Long Description','Default Cost','Original Price','Price','Primary Vendor','Color','Category','Size','Style Name','Style Number','Season','Taxable?','Sell on Shopify','Shopify Product Type','Shopify Tags','Shopify Category','Shopify Color','Active fit guide','Tags','Dimension 1','Dimension 1 Values','Dimension 2','Dimension 2 Values','Dimension 3','Dimension 3 Values','Active','Generate Items','Item #','Item Description','Item Default Cost','Item Tags','Item Dynamic Margin','Item Original Price','Item Current Price','Item Active?','Item Track Inventory?','Item Long Description','Item Prompt for description?','Item Prompt for price?','Item Primary Image','Item Use dynamic margin?','Item Primary Vendor','Item Color','Item Category','Item Size','Item Style Name','Item Style Number','Item Season','Item Taxable?','Item Sell on Shopify','Item Shopify Product Type','Item Shopify Tags','Item Shopify Category','Item Shopify Color','Item Active fit guide','Item Weight','Item Weight Unit','Item Width','Item Width Unit','Item Height','Item Height Unit','Item Depth','Item Depth Unit','Item romi boutique Reorder Point','Item romi boutique Target Qty','Item romi boutique Bin Location','Item Shopify Reorder Point','Item Shopify Target Qty','Item Shopify Bin Location','Item Max Retail Reorder Point','Item Max Retail Target Qty','Item Max Retail Bin Location','Item Garmentory Reorder Point','Item Garmentory Target Qty','Item Garmentory Bin Location','Item All Locations Reorder Point','Item All Locations Target Qty','Item All Locations Bin Location','Item Vendor Details Item #','Item Vendor Details Default Cost','Item Vendor Details #','Item Images URL']

# Non-size headers in the wide (Order-to-Size) line table. Anything else with a
# non-empty header between "Country of Origin" and the money columns is a size.
WIDE_NON_SIZE = {"style image", "style name", "style number", "color", "color code",
                 "color comment", "style comment", "materials", "fabrication",
                 "country of origin", "item discount", "units"}
WIDE_NON_SIZE_PREFIXES = ("sugg. retail", "sugg retail", "wholesale", "total")

# ------------------------- text utilities -----------------------------------
def title_caps(s: str) -> str:
    """Word-wise title case, capitalizing after spaces and hyphens,
    keeping letters after apostrophes lowercase (Isabel's, Off-White)."""
    def cap_word(w):
        parts = re.split(r"([-/])", w)   # keep the delimiters, capitalize each chunk
        return "".join(p[:1].upper() + p[1:].lower() if p and p not in "-/" else p
                      for p in parts)
    return " ".join(cap_word(w) for w in s.strip().split())

def fix_mojibake(s: str) -> str:
    """Repair UTF-8 text that was mis-decoded as Mac Roman / cp1252 upstream
    (e.g. 'D√¥en' -> 'Dôen'). Best effort; returns input if repair fails."""
    if not any(ch in s for ch in ("√", "Ã", "â", "€")):
        return s
    for enc in ("mac_roman", "cp1252"):
        try:
            fixed = s.encode(enc).decode("utf-8")
            if sum(ord(c) > 127 for c in fixed) < sum(ord(c) > 127 for c in s):
                return fixed
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
    return s

def clean_size(raw) -> str:
    """Strip a leading 'SIZE' label some vendors include in JOOR (e.g.
    'SIZE 32' -> '32', 'Size34' -> '34'), regardless of sizing convention.
    Applied wherever a raw size token first enters the pipeline."""
    s = str(raw).strip()
    return re.sub(r"(?i)^size\s*", "", s).strip()

def _vendor_key(s: str) -> str:
    """Normalize a vendor string for matching: strip accents/case/punctuation
    so 'Dôen', 'DOEN', and 'Do\u0308en' all collapse to the same key."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper().replace("&", "AND")
    s = re.sub(r"[.\',/]", " ", s)
    return re.sub(r"[^A-Z0-9]+", " ", s).strip()

_VENDOR_LOOKUP = {}
for _name in KNOWN_VENDOR_NAMES:
    _VENDOR_LOOKUP.setdefault(_vendor_key(_name), _name)

def normalize_vendor(raw: str) -> str:
    """Resolve raw JOOR vendor text to the EXACT existing Heartland vendor
    Name (matches how Heartland's import screen looks up vendors) rather than
    guessing a new string, which risks creating a duplicate vendor record."""
    raw = fix_mojibake(raw)
    if raw in VENDOR_MAP:
        return VENDOR_MAP[raw]
    name = raw.split(" (")[0]                                   # drop "(shopdoen.com)"
    hit = _VENDOR_LOOKUP.get(_vendor_key(name))
    if hit:
        return hit
    guess = unicodedata.normalize("NFKD", name)
    guess = "".join(c for c in guess if not unicodedata.combining(c)).upper().strip()
    print(f"  WARNING: vendor '{name}' not found in the known Heartland vendor "
          f"list; using '{guess}' as a guess. Check Purchasing -> Vendors before "
          f"importing, or add an override to VENDOR_MAP, to avoid creating a "
          f"duplicate vendor record.")
    return guess
    
# ------------------------- Heartland API (live dedupe check) ----------------
def _load_heartland_env():
    """Read HEARTLAND_TOKEN / HEARTLAND_SUBDOMAIN from a .env file next to
    this script, without requiring python-dotenv."""
    env_path = Path(__file__).resolve().parent / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.strip() and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

_load_heartland_env()
HEARTLAND_TOKEN = os.environ.get("HEARTLAND_TOKEN")
HEARTLAND_SUBDOMAIN = os.environ.get("HEARTLAND_SUBDOMAIN")
HEARTLAND_BASE_URL = (f"https://{HEARTLAND_SUBDOMAIN}.retail.heartland.us/api"
                      if HEARTLAND_SUBDOMAIN else None)
HEARTLAND_HEADERS = {"Authorization": f"Bearer {HEARTLAND_TOKEN}"} if HEARTLAND_TOKEN else {}

try:
    from vendor_id_map import VENDOR_ID_MAP   # generated by api_test_6.py
except ImportError:
    VENDOR_ID_MAP = {}
    print("  WARNING: vendor_id_map.py not found next to this script; "
          "Heartland live dedupe will be skipped (grids will always be "
          "generated, never checked against existing ones).")

_VENDOR_ID_LOOKUP = {}
for _vname, _vid in VENDOR_ID_MAP.items():
    _VENDOR_ID_LOOKUP.setdefault(_vendor_key(_vname), _vid)

def resolve_vendor_id(normalized_vendor_name: str):
    """Map an already-normalize_vendor()'d name to its Heartland internal
    vendor id via vendor_id_map.py. Returns None if unmapped (e.g. a vendor
    added in Heartland since vendor_id_map.py was last regenerated)."""
    return _VENDOR_ID_LOOKUP.get(_vendor_key(normalized_vendor_name))

def heartland_grid_exists(vendor_id, style_number, color):
    """Live-check whether an item grid already exists in Heartland for this
    vendor + style number + color (query pattern confirmed against real
    data: positive/negative controls both passed). Returns True/False, or
    None if the check couldn't be performed (missing credentials, network
    error, unmapped vendor) -- callers should treat None as 'include the
    grid' rather than silently dropping a potentially-new item."""
    if not (HEARTLAND_TOKEN and HEARTLAND_BASE_URL and vendor_id):
        return None
    try:
        resp = requests.get(
            f"{HEARTLAND_BASE_URL}/item_grids",
            headers=HEARTLAND_HEADERS,
            params={
                "~[item_primary_vendor_id]": vendor_id,
                "~[item_custom@style_number]": style_number,
                "~[item_custom@color]": color,
                "per_page": 1,
            },
            timeout=10,
        )
        if resp.status_code != 200:
            print(f"  WARNING: Heartland dedupe check failed (HTTP {resp.status_code}) "
                  f"for '{style_number}' / '{color}'; including grid as a precaution "
                  f"-- verify manually before importing.")
            return None
        return resp.json().get("total", 0) > 0
    except requests.RequestException as e:
        print(f"  WARNING: Heartland dedupe check errored ({e}) for "
              f"'{style_number}' / '{color}'; including grid as a precaution "
              f"-- verify manually before importing.")
        return None

# Which /api/items field carries the "Item #" Heartland's PO import expects.
# Vendors proved id != public_id (the UI's "#" column was public_id), so this
# is deliberately a single switch: flip to "public_id" if api_test_8 shows the
# UI value matches public_id instead of id.
ITEM_NUMBER_FIELD = "id"

def _item_number(it):
    v = it.get(ITEM_NUMBER_FIELD)
    return str(v).strip() if v not in (None, "") else ""

def _item_custom(it):
    """Items nest custom fields under 'custom'; item_grids uses 'item_custom'.
    Accept either so a schema surprise doesn't silently produce an empty index."""
    for k in ("custom", "item_custom"):
        v = it.get(k)
        if isinstance(v, dict):
            return v
    return {}

_ITEM_INDEX_CACHE = {}

def clear_item_index_cache():
    """Drop cached vendor item indexes so the next lookup refetches. Called at
    the start of each build_po run: within a run the cache saves one request
    per JOOR line, but between runs new items may have been generated in
    Heartland (e.g. by the grids import that just happened), and a stale index
    would silently push those lines into the 'missing' warning."""
    _ITEM_INDEX_CACHE.clear()

def heartland_vendor_item_index(vendor_id, refresh=False):
    """One paginated fetch of every item for a vendor ->
    {(STYLE, COLOR, SIZE): item_no}. Replaces the per-JOOR-line lookup (one
    HTTP request per line) with ~one request per 100 items. Returns None if
    the fetch couldn't be completed, so build_po can stop rather than emit a
    PO that's silently missing lines."""
    if not (HEARTLAND_TOKEN and HEARTLAND_BASE_URL and vendor_id):
        return None
    if not refresh and vendor_id in _ITEM_INDEX_CACHE:
        return _ITEM_INDEX_CACHE[vendor_id]
    index, page = {}, 1
    while True:
        try:
            resp = requests.get(
                f"{HEARTLAND_BASE_URL}/items",
                headers=HEARTLAND_HEADERS,
                params={"~[primary_vendor_id]": vendor_id,
                        "per_page": 100, "page": page},
                timeout=30,
            )
        except requests.RequestException as e:
            print(f"  ERROR: Heartland item fetch failed on page {page} ({e}).")
            return None
        if resp.status_code != 200:
            print(f"  ERROR: Heartland item fetch failed on page {page} "
                  f"(HTTP {resp.status_code}).")
            return None
        data = resp.json()
        for it in data.get("results", []):
            num = _item_number(it)
            if not num:
                continue
            cust = _item_custom(it)
            style = str(cust.get("style_number") or "").strip().upper()
            color = str(cust.get("color") or "").strip().upper()
            size = canonical_size_token(cust.get("size") or "").upper()
            if style and color:
                index[(style, color, size)] = num
        pages = data.get("pages", 1)
        print(f"  fetched item page {page}/{pages} ({len(index)} indexed)")
        if page >= pages:
            break
        page += 1
    _ITEM_INDEX_CACHE[vendor_id] = index
    return index

def heartland_find_item(vendor_id, style_number, color, size):
    """Live-look up the Heartland Item # for a vendor + style number + color +
    size via /api/items (NOTE: this endpoint uses plain 'custom@' and
    'primary_vendor_id', unlike item_grids' 'item_custom@'/'item_primary_
    vendor_id' -- confirmed via api_test_7.py). Returns the item id as a
    string if found, else None (missing credentials, network error, unmapped
    vendor, or genuinely no match) -- callers treat None as 'no match'."""
    if not (HEARTLAND_TOKEN and HEARTLAND_BASE_URL and vendor_id):
        return None
    try:
        resp = requests.get(
            f"{HEARTLAND_BASE_URL}/items",
            headers=HEARTLAND_HEADERS,
            params={
                "~[primary_vendor_id]": vendor_id,
                "~[custom@style_number]": style_number,
                "~[custom@color]": color,
                "~[custom@size]": size,
                "per_page": 1,
            },
            timeout=10,
        )
        if resp.status_code != 200:
            print(f"  WARNING: Heartland item lookup failed (HTTP {resp.status_code}) "
                  f"for '{style_number}' / '{color}' / '{size}'.")
            return None
        results = resp.json().get("results", [])
        return str(results[0]["id"]) if results else None
    except requests.RequestException as e:
        print(f"  WARNING: Heartland item lookup errored ({e}) for "
              f"'{style_number}' / '{color}' / '{size}'.")
        return None

def color_code(color: str) -> str:
    words = re.split(r"[\s\-]+", color.strip())
    return "".join(w[0] for w in words if w and w[0].isalnum()).upper()

def slug(s: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "-", s.upper()).strip("-")

def _num(v):
    """'1' / 1 / 1.0 / '' -> float or None."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None

# ------------------------- size completion engine ---------------------------
LETTER_LADDER = ["XXXS","XXS","XS","S","M","L","XL","XXL","XXXL","XXXXL"]
LETTER_BASE = ("XS","L")
OS_TOKENS = {"OS","O/S","ONE SIZE","ONESIZE"}

# Apparel numeric-even ladder (dresses/tops/etc): '00' is the rung one step
# below size 0 (NOT the same as '0'); the run commonly extends above 10 to
# 12/14/16/18+. Kept as strings throughout so '00' stays distinct from '0'.
NUMERIC_EVEN_LADDER = ["00","0","2","4","6","8","10","12","14","16","18","20"]

def _is_numeric_even_token(s):
    s = str(s).strip()
    if s == "00":
        return True
    try:
        v = int(s)
    except ValueError:
        return False
    return v >= 0 and v % 2 == 0

def _numeric_even_ladder_upto(sizes):
    """Extend the base 00..20 ladder further if an order runs above it."""
    ladder = list(NUMERIC_EVEN_LADDER)
    hi_val = max(-1 if str(s).strip() == "00" else int(s) for s in sizes)
    last_val = int(ladder[-1])
    while last_val < hi_val:
        last_val += 2
        ladder.append(str(last_val))
    return ladder

def detect_convention(sizes):
    """Return (name, ladder, base_lo_idx, base_hi_idx) or fallback."""
    up = [s.upper() for s in sizes]
    if all(s in OS_TOKENS for s in up):
        return ("os", ["OS"], 0, 0)
    if all(s in LETTER_LADDER for s in up):
        return ("letters", LETTER_LADDER,
                LETTER_LADDER.index(LETTER_BASE[0]), LETTER_LADDER.index(LETTER_BASE[1]))
    if all(_is_numeric_even_token(s) for s in sizes):
        # Handles plain runs (2,4,6,8), sizes that dip to '00', and runs that
        # climb past 10 (12,14...) -- all as one continuous even ladder.
        ladder = _numeric_even_ladder_upto(sizes)
        return ("numeric-even", ladder, ladder.index("0"), ladder.index("10"))
    try:
        nums = sorted({int(s) for s in sizes})
    except ValueError:
        return ("unknown", None, None, None)
    lo, hi = nums[0], nums[-1]
    if hi <= 12:
        odd = any(n % 2 for n in nums)
        if odd and hi <= 6:                       # 0,1,2,3,4 style
            ladder = list(range(0, max(hi, 4) + 1))
            return ("numeric-0-4", ladder, ladder.index(0), ladder.index(4))
        return ("unknown", None, None, None)       # odd sizes above 6: unrecognized
    if 20 <= lo and hi <= 45:                      # denim 24-29 base, step 1
        ladder = list(range(min(lo, 24), max(hi, 29) + 1))
        return ("denim-24-29", ladder, ladder.index(24), ladder.index(29))
    if lo >= 60 and all(n % 5 == 0 for n in nums): # euro 70-90 base, step 5
        ladder = list(range(min(lo, 70), max(hi, 90) + 5, 5))
        return ("euro-70-90", ladder, ladder.index(70), ladder.index(90))
    return ("unknown", None, None, None)

def complete_sizes(ordered_sizes):
    """Continuous ladder span from min(base, ordered) to max(base, ordered)."""
    name, ladder, b_lo, b_hi = detect_convention(ordered_sizes)
    if name == "os":
        return ["OS"], name
    if name == "unknown":
        print(f"  WARNING: unrecognized size convention {sorted(set(ordered_sizes))}; "
              f"using ordered sizes as-is (no completion).")
        return sorted(set(str(s) for s in ordered_sizes)), name
    if name == "letters":
        idxs = [ladder.index(s.upper()) for s in ordered_sizes]
    elif name == "numeric-even":
        idxs = [ladder.index("00" if str(s).strip() == "00" else str(int(s)))
                for s in ordered_sizes]
    else:
        idxs = [ladder.index(int(s)) for s in ordered_sizes]
    lo, hi = min(min(idxs), b_lo), max(max(idxs), b_hi)
    return [str(ladder[i]) for i in range(lo, hi + 1)], name

def canonical_size_token(raw) -> str:
    """Normalize a single raw JOOR size token to the exact string Heartland
    stores as an item's size after grid generation (same rules build_grids
    applies via detect_convention/complete_sizes, but for one token instead
    of filling a range) -- used by build_po so its live item lookup always
    queries with the same token build_grids used to create the item."""
    s = clean_size(raw)
    su = s.upper()
    if su in OS_TOKENS:
        return "OS"
    if su in LETTER_LADDER:
        return su
    if su == "00":
        return "00"
    try:
        return str(int(su))
    except ValueError:
        return s   # unrecognized convention: build_grids keeps original casing too

def _numeric_even_span(lo_token, hi_token):
    """Inclusive even-step run between two numeric-even tokens, e.g. '00'..'14'."""
    ladder = _numeric_even_ladder_upto([lo_token, hi_token])
    lo_key = "00" if lo_token == "00" else str(int(lo_token))
    hi_key = "00" if hi_token == "00" else str(int(hi_token))
    return ladder[ladder.index(lo_key):ladder.index(hi_key) + 1]

def expand_size_ranges(rows):
    """JOOR Basic Export sometimes rolls a whole run of sizes into a single
    row, e.g. size='00-14' with quantity=2, meaning 2 units of EACH even size
    from 00 (the rung below size 0) through 14 -- not 2 units total. Expand
    each such row into one row per size, preserving the per-size quantity."""
    out = []
    for r in rows:
        size = clean_size(r.get("size", ""))
        m = re.fullmatch(r"(\d{1,2})-(\d{1,2})", size)
        if not m:
            nr = dict(r); nr["size"] = size
            out.append(nr)
            continue
        lo, hi = m.groups()
        if not (_is_numeric_even_token(lo) and _is_numeric_even_token(hi)):
            nr = dict(r); nr["size"] = size   # not a run we recognize; leave cleaned
            out.append(nr)
            continue
        sizes = _numeric_even_span(lo, hi)
        print(f"  NOTE: '{r.get('style_name','')}' size '{size}' read as a run "
              f"-> expanded to {sizes} ({r.get('quantity')} unit(s) each)")
        for s in sizes:
            nr = dict(r); nr["size"] = s
            out.append(nr)
    return out

# ------------------------- pricing -------------------------------------------
def ceil_dollars(val) -> int:
    return int(Decimal(str(val)).to_integral_value(rounding=ROUND_CEILING))

def retail_price(unit_cost, sugg_retail=None) -> int:
    """JOOR Sugg. Retail if present, else MARGIN x cost; always ceil to whole $."""
    if sugg_retail:
        return ceil_dollars(sugg_retail)
    return ceil_dollars(Decimal(str(unit_cost)) * MARGIN)

# ------------------------- io: format detection & parsing --------------------
def _load_sheet_rows(path):
    """xlsx -> (worksheet, list-of-list of string cell values). data_only skips formulas."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = [[("" if c is None else str(c).strip()) for c in r]
            for r in ws.iter_rows(values_only=True)]
    return ws, rows

def _find_meta(rows, label):
    """Find 'Label:' either as its own cell (value = next non-empty cell to the
    right) or embedded 'Label: value' inside one cell."""
    lab = label.lower().rstrip(":")
    for r in rows:
        for i, cell in enumerate(r):
            low = cell.lower()
            if low.rstrip(":") == lab and cell.endswith(":"):
                for nxt in r[i+1:]:
                    if nxt:
                        return nxt
            m = re.match(rf"{re.escape(lab)}\s*:\s*(.+)", low)
            if m:
                return cell.split(":", 1)[1].strip()
    return ""

def _looks_like_color_code(s: str) -> bool:
    """True for short alphanumeric codes like 'C0308' or 'R0272' (no spaces,
    mixes letters+digits, short) as opposed to a real color name like
    'Oatmeal Melange Multi' or 'Cardinal Red'."""
    s = s.strip()
    if not s:
        return True                     # empty -> not a usable name either
    if " " in s or "/" in s:
        return False                    # multi-word or slash-combo -> real name
    return bool(re.fullmatch(r"[A-Za-z]{0,3}\d+[A-Za-z0-9]*", s))

def resolve_color(color_val, color_code_val):
    """JOOR's 'Color' vs 'Color Code' columns hold the descriptive name and the
    internal code in either order depending on the vendor/brand (confirmed:
    DOEN puts the name in 'Color'; Veronica Beard puts it in 'Color Code').
    Pick whichever cell actually looks like a name; fall back to 'Color' if
    both or neither look like one."""
    color_val, color_code_val = (color_val or "").strip(), (color_code_val or "").strip()
    if color_val and not _looks_like_color_code(color_val):
        return color_val
    if color_code_val and not _looks_like_color_code(color_code_val):
        return color_code_val
    return color_val or color_code_val  # both look like codes (or one's blank): best guess

def parse_order_to_size(path):
    """JOOR 'Order to Size' wide xlsx -> (normalized long rows, images dict)."""
    _, rows = _load_sheet_rows(path)

    # --- header row of the line-item table
    hdr_idx = next((i for i, r in enumerate(rows)
                    if "style number" in [c.lower() for c in r]
                    and "style name" in [c.lower() for c in r]), None)
    if hdr_idx is None:
        sys.exit("ERROR: couldn't find the line-item header row (Style Name / Style Number).")
    hdr = rows[hdr_idx]
    col = {c.lower(): i for i, c in enumerate(hdr) if c}
    size_cols = [(i, c) for i, c in enumerate(hdr) if c
                 and c.lower() not in WIDE_NON_SIZE
                 and not c.lower().startswith(WIDE_NON_SIZE_PREFIXES)]
    ws_col = next((i for c, i in col.items() if c.startswith("wholesale")), None)
    sr_col = next((i for c, i in col.items() if c.startswith(("sugg. retail", "sugg retail"))), None)
    if ws_col is None:
        sys.exit("ERROR: no WholeSale column found in the line-item table.")

    # --- order-level metadata
    vendor_raw = next((c for r in rows[:hdr_idx] for c in r if c), "")
    if vendor_raw.endswith(":") or not vendor_raw:
        vendor_raw = ""
        print("  WARNING: couldn't identify the vendor cell; set VENDOR_MAP or check output.")
    order_no  = _find_meta(rows[:hdr_idx], "PO Number")
    start_ship = _find_meta(rows[:hdr_idx], "Start Ship")
    end_ship   = _find_meta(rows[:hdr_idx], "Complete")
    print(f"  [order-to-size] vendor='{vendor_raw}'  PO={order_no}  "
          f"ship {start_ship} -> {end_ship}  (no linesheet column: Season will be blank)")

    # --- melt wide line rows into the same long dicts the pipeline expects
    long_rows, line_row_idx = [], {}
    for i in range(hdr_idx + 1, len(rows)):
        r = rows[i]
        style_no  = r[col["style number"]] if col["style number"] < len(r) else ""
        style_name = r[col["style name"]] if col["style name"] < len(r) else ""
        if not style_no or not style_name:
            continue                                   # totals / footer rows
        color_raw  = r[col["color"]] if col["color"] < len(r) else ""
        code_col_idx = col.get("color code")
        color_code_raw = r[code_col_idx] if code_col_idx is not None and code_col_idx < len(r) else ""
        color = resolve_color(color_raw, color_code_raw)
        if color and color != color_raw.strip():
            print(f"  NOTE: '{style_name}': using Color Code column value "
                  f"'{color}' as the real color name ('{color_raw}' in the "
                  f"Color column looks like an internal code for this vendor).")
        line_row_idx[i] = (style_no, color)            # for image anchoring
        for ci, raw_label in size_cols:
            qty = _num(r[ci]) if ci < len(r) else None
            if not qty:
                continue
            size_label = clean_size(raw_label)
            m = re.fullmatch(r"(\d{1,2})-(\d{1,2})", size_label)
            if m and _is_numeric_even_token(m.group(1)) and _is_numeric_even_token(m.group(2)):
                run_sizes = _numeric_even_span(*m.groups())
                print(f"  NOTE: '{style_name}' size column '{size_label}' read as a "
                      f"run -> expanded to {run_sizes} ({qty} unit(s) each)")
            else:
                run_sizes = [size_label]
            for size_val in run_sizes:
                long_rows.append({
                    "seller_account": vendor_raw,
                    "style_number":  style_no,
                    "style_name":    style_name,
                    "color":         color,
                    "size":          size_val,
                    "quantity":      qty,
                    "item_price":    _num(r[ws_col]) or 0,
                    "sugg_retail":   _num(r[sr_col]) if sr_col is not None else None,
                    "order_number":  order_no,
                    "start_date":    start_ship,
                    "cancel_date":   end_ship,
                    SEASON_FROM:     "",
                })
    return long_rows

def read_joor(path):
    """Read a JOOR export in either format -> (list of long dicts, images dict).
    Auto-detects: Basic Export (long) vs Order-to-Size (wide, formatted xlsx)."""
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        _, rows = _load_sheet_rows(p)
        first_hdr = [h.lower() for h in rows[0]] if rows else []
        if "seller_account" not in first_hdr:                    # wide format
            return parse_order_to_size(p)
        rows = [[c for c in r] for r in rows]
    else:
        raw = p.read_bytes()
        for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = raw.decode(enc); break
            except UnicodeDecodeError:
                continue
        delim = "\t" if text.splitlines()[0].count("\t") > text.splitlines()[0].count(",") else ","
        rows = list(csv.reader(text.splitlines(), delimiter=delim))
    hdr = [h.strip().lower() for h in rows[0]]
    long_rows = [dict(zip(hdr, r)) for r in rows[1:] if any(str(c).strip() for c in r)]
    return expand_size_ranges(long_rows), {}

def write_xlsx(path, headers, rows, text_cols=(), dropdowns=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = Workbook(); ws = wb.active
    bold = Font(name="Arial", size=10, bold=True); norm = Font(name="Arial", size=10)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = bold
    for r_i, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            v = row.get(h)
            cell = ws.cell(row=r_i, column=c)
            if h in text_cols and v is not None:
                cell.value = str(v); cell.number_format = "@"
            else:
                cell.value = v
            cell.font = norm
    for c, h in enumerate(headers, 1):
        vals = [h] + [str(r.get(h) or "") for r in rows]
        ws.column_dimensions[get_column_letter(c)].width = min(max(len(s) for s in vals) + 2, 34)
    ws.freeze_panes = "A2"

    if dropdowns:
        # Hidden helper sheet holding each option list; data validation formulas
        # reference it by range (more reliable across Excel/Numbers than long
        # inline comma lists, which also risk hitting the ~255-char limit).
        lists_ws = wb.create_sheet("Lists")
        lists_ws.sheet_state = "hidden"
        last_row = max(len(rows) + 100, 200)   # room for manually added rows later
        for col_i, (header, options) in enumerate(dropdowns.items(), 1):
            if header not in headers:
                continue
            target_col_idx = headers.index(header) + 1
            col_letter = get_column_letter(target_col_idx)  # use target, not col_i
            for r_i, opt in enumerate(options, 1):
                lists_ws.cell(row=r_i, column=target_col_idx, value=opt)
            dv = DataValidation(
                type="list",
                formula1=f"=Lists!${col_letter}$1:${col_letter}${len(options)}",
                allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{last_row}")

    wb.save(path)

# ------------------------- phase 1: grids ------------------------------------
def build_grids(joor_rows, out_path):
    groups = {}                                   # (vendor, style#, color) -> data
    for r in joor_rows:
        vendor = normalize_vendor(r["seller_account"])
        raw_color = r["color"]
        if _looks_like_color_code(raw_color):
            print(f"  NOTE: color '{raw_color}' for '{r['style_name']}' looks like a "
                  f"vendor code, not a color name. This export format doesn't carry a "
                  f"plain-English color column -> re-download using JOOR's 'Download "
                  f"Order Summary Excel' option instead, which does.")
        key = (vendor, r["style_number"].strip(), title_caps(fix_mojibake(raw_color)))
        g = groups.setdefault(key, {"sizes": [], "cost": r["item_price"],
                                    "sugg_retail": r.get("sugg_retail"),
                                    "style_name_raw": r["style_name"],
                                    "raw_key": (r["style_number"].strip(), r["color"]),
                                    "season": str(r.get(SEASON_FROM, "") or "").strip()})
        g["sizes"].append(str(r["size"]).strip())

    # grid numbers with intra-run collision handling
    codes = {}
    for (vendor, style_no, color) in groups:
        code = color_code(color)
        while (vendor, style_no, code) in codes.values():
            code = slug(color)                    # fall back to full color slug
        codes[(vendor, style_no, color)] = (vendor, style_no, code)

    out_rows, skipped_keys = [], set()
    for key, g in groups.items():
        vendor, style_no, color = key
        vendor_id = resolve_vendor_id(vendor)
        if heartland_grid_exists(vendor_id, style_no, color):
            skipped_keys.add(key)
            print(f"  SKIP (already exists in Heartland): {style_no} / {color} ({vendor})")
            continue
        desc = title_caps(fix_mojibake(g["style_name_raw"]))
        sizes, conv = complete_sizes(g["sizes"])
        price = retail_price(g["cost"], g.get("sugg_retail"))
        src = "JOOR Sugg. Retail" if g.get("sugg_retail") else f"{MARGIN}x cost"
        print(f"  {desc} / {color}: ordered {sorted(set(g['sizes']))} -> "
              f"[{conv}] generated {sizes}  | price ${price} ({src})")
        if not g["season"]:
            print(f"    NOTE: no Season available for this line; fill in manually.")
        row = {h: None for h in TEMPLATE_HEADERS}
        row.update({
            "Item Grid #": "-".join(codes[key]),
            "Description": desc,
            "Default Cost": float(g["cost"]),
            "Price": price,
            "Primary Vendor": vendor,
            "Style Name": f"{desc} in {color}",
            "Style Number": style_no,
            "Season": g["season"] or None,
            "Dimension 1": "Color", "Dimension 1 Values": color,
            "Dimension 2": "Size", "Dimension 2 Values": ",".join(sizes),
            "Active": "true", "Generate Items": "true",
            "Active fit guide": "3",
        })
        out_rows.append(row)
    write_xlsx(out_path, TEMPLATE_HEADERS, out_rows,
               text_cols=("Item Grid #", "Style Number"), dropdowns=DROPDOWN_LISTS)
    print(f"wrote {out_path} ({len(out_rows)} grid(s))")
    if skipped_keys:
        print(f"  skipped {len(skipped_keys)} grid(s) already existing in Heartland")

# ------------------------- phase 2: PO ---------------------------------------
PO_HEADERS = ["PO #", "PO Description", "PO Vendor", "PO Start Ship", "PO End Ship", "PO Received at location",
              "Item #", "PO Line Qty", "PO Line Unit Cost", "Item Default Cost"]

def po_description(order_number, shipping_fee=None):
    """PO Description convention: '<order_number>' normally, or
    'F:$<fee>/<order_number>' when a per-order shipping fee was charged."""
    if shipping_fee:
        return f"F:${ceil_dollars(shipping_fee)}/{order_number}"
    return str(order_number)
# PO # placeholder is used because it is a required field upon PO import.
# Heartland doesn't allow auto-generation of the PO # upon PO import, so this
# placeholder allows the import to happen but requires user follow up. The JOOR
# order number goes in PO Description instead, so it's still visible/searchable
# without overriding Heartland's numbering.
RECEIVED_AT_LOCATION = "romi boutique"   # Heartland requires this field on PO import; single-location store
# Helper columns are parenthesized on purpose: Heartland won't auto-map them,
# so they import as Ignored and can't overwrite anything.

def _size_sort_key(size):
    s = str(size).strip().upper()
    if s in LETTER_LADDER:
        return (0, LETTER_LADDER.index(s), "")
    try:
        return (1, float(s), "")
    except ValueError:
        return (2, 0, s)

def _iget(it, *keys):
    """First non-empty value among several possible header names. Heartland's
    ITEMS export uses unprefixed headers (Description, Color, Size, Style
    Number...) while its GRIDS export prefixes them with 'Item '. Accept both."""
    for k in keys:
        v = it.get(k)
        if v not in (None, ""):
            return str(v)
    return ""

def build_po(joor_rows, out_path):
    clear_item_index_cache()   # items may have been generated since the last run
    out_rows, missing = [], []
    for r in joor_rows:
        vendor = normalize_vendor(r["seller_account"])
        vendor_id = resolve_vendor_id(vendor)
        index = heartland_vendor_item_index(vendor_id)
        if index is None:
            sys.exit(f"ERROR: couldn't load Heartland items for vendor '{vendor}'.\n"
                     "  -> check the .env credentials and network, and that this "
                     "vendor exists in vendor_id_map.py. No PO file written.")
        color = title_caps(fix_mojibake(r["color"]))
        size = canonical_size_token(r["size"])
        item_no = index.get((r["style_number"].strip().upper(),
                             color.upper(), size.upper()))
        if not item_no:
            missing.append((r["style_number"], color, r["size"])); continue
        cost = float(r["item_price"])
        out_rows.append({
            "PO #": "CHECK HEARTLAND HIGHEST PO#"
            "PO Description": po_description(r["order_number"], r.get("shipping_fee")),
            "PO Vendor": vendor,
            "PO Start Ship": r.get("start_date", ""), "PO End Ship": r.get("cancel_date", ""),
            "PO Received at location": RECEIVED_AT_LOCATION,
            "Item #": item_no, "PO Line Qty": int(float(r["quantity"])),
            "PO Line Unit Cost": cost, "Item Default Cost": cost,
        })
    if missing:
        print(f"  WARNING: {len(missing)} JOOR line(s) had no matching Heartland item "
              f"(vendor/style/color/size): {missing}\n  -> the vendor's items WERE "
              f"loaded, so these specific style/color/size combos aren't in Heartland: "
              f"check the grids import ran and generated items for every size.")
    write_xlsx(out_path, PO_HEADERS, out_rows, text_cols=("Item #",))
    print(f"wrote {out_path} ({len(out_rows)} PO line(s))")

# ------------------------- cli ------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="JOOR -> Heartland converter (Romi Boutique)")
    sub = ap.add_subparsers(dest="cmd", required=True)
    g = sub.add_parser("grids"); g.add_argument("joor_file"); g.add_argument("-o", default="grids_import.xlsx")
    p = sub.add_parser("po"); p.add_argument("joor_file")
    p.add_argument("-o", default="po_import.xlsx")
    a = ap.parse_args()
    rows, images = read_joor(a.joor_file)
    print(f"read {len(rows)} JOOR line(s) from {a.joor_file}")
    if a.cmd == "grids":
        build_grids(rows, a.o)
    else:
        build_po(rows, a.o)

if __name__ == "__main__":
    main()
